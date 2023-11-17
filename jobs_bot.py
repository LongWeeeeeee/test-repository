from openpyxl import load_workbook
from datetime import datetime, timedelta
import telebot
import requests
import json
from numbers_parser import Document
bot = telebot.TeleBot(token='6635829285:AAGhpvRdh-6DtnT6DveZEky0tt5U_PejLXs')
def send_message(message):
    BOT_TOKEN = '6635829285:AAGhpvRdh-6DtnT6DveZEky0tt5U_PejLXs'
    CHAT_ID = '1091698279'
    url = f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage'
    payload = {
        'chat_id': CHAT_ID,
        'text': message
    }
    response = requests.post(url, json=payload)
    result = json.loads(response.content)
    return result

def add_day_to_excel(date, activities, user_message, total_sleep, deep_sleep, rate, mysteps):
    # Загрузка существующего файла Excel
    wb = load_workbook('MyDays.xlsx')

    # Получение активного листа
    sheet = wb.active

    # Поиск последней заполненной строки в столбце "Дата"
    last_row = sheet.max_row

    # Получение вчерашней даты
    yesterday = date - timedelta(days=1)

    # Вставка вчерашней даты в следующую за последней заполненной строкой
    sheet.cell(row=last_row+1, column=1).value = yesterday.strftime("%d.%m.%Y")

    # Вставка дел в таблицу
    sheet.cell(row=last_row+1, column=2).value = ", ".join(activities)
    # Вставка шаги
    sheet.cell(row=last_row + 1, column=3).value = mysteps
    # Вставка сколько часов спал
    sheet.cell(row=last_row + 1, column=4).value = total_sleep
    # Вставка времени глубокого сна
    sheet.cell(row=last_row + 1, column=5).value = deep_sleep
    # Вставка о дне в таблицу
    sheet.cell(row=last_row + 1, column=6).value = user_message
    # Вставка времени глубокого сна
    sheet.cell(row=last_row + 1, column=7).value = rate

    # Вычисление итогового счета
    score = 0
    for activity in activities:
        score += scores[activity]

    # Вставка итогового счета в таблицу
    sheet.cell(row=last_row+1, column=8).value = score

    # Сохранение изменений в файл
    wb.save("MyDays.xlsx")
    send_message('Готово! Хорошего тебе дня, пусть он будет лучше, чем вчера :)')


scores = {'встал в 6:30': 1, 'лег в 11': 1, 'умылся льдом': 1, 'контрастный душ': 1, 'сделал зарядку': 1, 'дрочил': -1, 'позанимался вокалом' : 1, 'сходил за водой': 1,'правильно питался': 1, 'читал книгу': 1, 'шаги': 1, 'принимал витамины': 1, 'массаж перед сном': 1, 'сахар': -1}

send_message('Расскажи мне как провел вчерашний день?' + '\n' + 'Вот возможный список дел:' + '\n' + '\n' +  ', '.join(scores.keys()))

@bot.message_handler(content_types=['text'])
def send_text(message):
    global activities
    activities = []
    flag = False
    user_message=message.text
    for activy in user_message.split(', '):
        if activy in scores:
            flag = True
            activities.append(activy)
        else:
            send_message('Активности ' + user_message + ' нету в списке...')
    if flag:
        msg = bot.reply_to(message,
                           'Хочешь рассказать как прошел день? Это поможет отслеживать почему день был хороший или нет')
        bot.register_next_step_handler(msg, process_name_step)

def process_name_step(message):
    global user_message
    user_message = message.text
    msg = bot.reply_to(message,
                       'Сколько сделал шагов?')
    bot.register_next_step_handler(msg, steps)
def steps(message):
    global mysteps
    mysteps = message.text
    msg = bot.reply_to(message,
                       'Сколько всего спал?')
    bot.register_next_step_handler(msg, process_sleep_total)
def process_sleep_total(message):
    global total_sleep
    total_sleep = message.text
    msg = bot.reply_to(message,
                       'Сколько из них глубокий сон?')
    bot.register_next_step_handler(msg, process_deep_sleep)


def process_deep_sleep(message):
    global deep_sleep
    deep_sleep = message.text
    msg = bot.reply_to(message,
                       'Насколько из 10 сам оцениваешь день?')
    bot.register_next_step_handler(msg, my_rate)
def my_rate(message):
    global rate
    rate = message.text
    date = datetime.now()
    add_day_to_excel(date, activities, user_message, total_sleep, deep_sleep, rate, mysteps)


bot.polling()