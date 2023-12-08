import asyncio
from sqlite import db_start, create_profile, edit_database
from openpyxl import load_workbook
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import Message
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from aiogram import types
from aiogram import F
import copy
from aiogram.filters import Command
import pandas as pd
from aiogram.types import FSInputFile
import json
import os
from openpyxl.workbook import Workbook
import keys

bot = Bot(token=keys.Token)
dp = Dispatcher()
start = True

class ClientState(StatesGroup):
    greet = State()
    new_daily_scores = State()
    steps = State()
    total_sleep = State()
    deep_sleep = State()
    about_day = State()
    personal_rate = State()
    settings = State()
    download = State()
    jobs = State()
    one_time_jobs_2 = State()
    one_time_jobs_proceed = State()
    date_jobs = State()
    date_jobs_2 = State()
async def on_startup():
    await db_start()

@dp.message(Command(commands=["start"]))
async def greetings(message: Message, state: FSMContext):
    # await create_profile(user_id=message.from_user.id)
    answer = await create_profile(user_id=message.from_user.id)
    if answer is not None:
        if answer[1] != '':
            daily_scores = json.loads(answer[1])
            await state.update_data(daily_scores=daily_scores)
        if answer[2] != '':
            one_time_jobs = json.loads(answer[2])
            await state.update_data(one_time_jobs=one_time_jobs)
        if answer[3] != '':
            date_jobs = json.loads(answer[3])
            await state.update_data(date_jobs=date_jobs)
        print(answer)
        await existing_user(message, state)
    else:
        await handle_new_user(message, state)
    start_scheduler(message, state)


@dp.message(F.text == 'Вывести дневник')
async def settings(message: Message, state: FSMContext):
    await diary_out(message)
    await state.set_state(ClientState.greet)


@dp.message(F.text == 'Изменить Настройки')
async def settings(message: Message, state: FSMContext):
    keyboard = create_settings_keyboard()
    await message.answer(text='Здесь вы можете изменить свои настройки', reply_markup=keyboard)
    await state.set_state(ClientState.settings)


@dp.message(F.text == 'Заполнить дневник', ClientState.settings)
async def fill_diary(message: Message, state: FSMContext) -> None:
    await greetings(message, state)


@dp.message(F.text == 'Изменить Стоимость', ClientState.settings)
async def download(message: Message, state: FSMContext = None) -> None:
    await message.answer(
        'Ниже представлен ваш ежедневный список дел и его стоимость в формате "дело: оценка" '
        'Скопируйте список и отправьте его с обновленными оценками', reply_markup=types.ReplyKeyboardRemove())
    user_states_data = await state.get_data()
    daily_scores = user_states_data['daily_scores']
    formatted_string = ''
    for key, value in daily_scores.items():
        formatted_string += f"{key} : {value}, "
    await message.answer(formatted_string[:-2])
    await state.set_state(ClientState.new_daily_scores)


@dp.message(F.text == '''Что такое "стоимость"?''', ClientState.settings)
async def download(message: Message) -> None:
    await message.answer(
        '''Стоимость нужна для подсчета очков за ваш день. Изначально стоимость любого дела равна 1,
но вы можете переназначить собственное значение. Нужны они для того составить обьективную картину.\n
Представим ситуацию, у вас был тяжелый день, и вы чувтсвуете себя совершенно разбитым
под конец дня и ставите этому дню оценку 0, однако за этот же день вы делали что-то из
списка своих дел, например правильно питались, учили новые языки или еще что-то, и день
уже не кажется таким плохим, потому что вы сделали свою рутину. Очки это обьективная оценка
дня, тогда как ваша персональная оценка это субьективная оценка и может быть не до конца точной''')


@dp.message(F.text == 'Изменить Дела', ClientState.settings)
async def jobs_change(message: Message, state: FSMContext = None) -> None:
    keyboard = create_jobs_keyboard()
    await message.answer(text='Вывожу список...', reply_markup=keyboard)
    await state.set_state(ClientState.jobs)


@dp.message(F.text == 'В определенную дату', ClientState.jobs)
async def date_jobs(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    if 'date_jobs' in data:
        await message.answer('Вот ваш список будующих дел:')
        output = "\n".join([f"{key} : {value}" for key, value in data['date_jobs'].items()])
        await message.answer(output)
        await message.answer(
            'Введите дело о котором нужно будет напомнить в определенную дату и оно добавится к списку', reply_markup=types.ReplyKeyboardRemove())
    else:
        await message.answer('Введите дело о котором нужно будет напомнить в определенную дату')
    await state.set_state(ClientState.date_jobs)


@dp.message(ClientState.date_jobs)
async def date_jobs(message: Message, state: FSMContext) -> None:
    await state.update_data(new_date_jobs=message.text)
    await message.answer('Введите дату когда вам о нем напомнить в формате год-месяц-день, например:')
    await message.answer(datetime.now().strftime("%Y-%m-%d"))
    await state.set_state(ClientState.date_jobs_2)


@dp.message(ClientState.date_jobs_2)
async def date_jobs(message: Message, state: FSMContext) -> None:
    user_states_data = await state.get_data()
    new_date_jobs = user_states_data['new_date_jobs']
    if 'date_jobs' in user_states_data:
        date_jobs = user_states_data['date_jobs']
        date_jobs[new_date_jobs] = message.text
    else:
        date_jobs = {new_date_jobs : message.text}
    await edit_database(date_jobs=date_jobs)
    await message.answer('Дело добавлено!')
    await state.update_data(date_jobs=date_jobs)
    await state.set_state(ClientState.settings)
    await settings(message, state)
    # with open('daily_scores.txt', 'r+', encoding='utf-8-sig') as f:
    #     data = json.load(f)
    #     data[str(message.from_user.id)]['date_jobs'][date_jobs] = message.text
    #     write_to_file(data, f)
    #
    #     dict = [key + ' : ' + value for key, value in data[str(message.from_user.id)]['date_jobs'].items()]
    #     await state.update_data(date_jobs=dict)
    #     await state.set_state(ClientState.settings)
    #     await settings(message, state)


@dp.message(F.text == 'Разовые дела', ClientState.jobs)
async def one_time_jobs(message: Message, state: FSMContext) -> None:
    user_states_data = await state.get_data()
    if 'one_time_jobs' in user_states_data:
        one_time_jobs = user_states_data['one_time_jobs']
        await message.answer(
            'Введите новый список разовых дел через запятую. Ваш предыдущий список:',
            reply_markup=types.ReplyKeyboardRemove())
        await message.answer(', '.join(one_time_jobs))
        await state.set_state(ClientState.one_time_jobs_2)
    else:
        await message.answer('Введите новый список разовых дел через запятую', reply_markup=types.ReplyKeyboardRemove())
        await state.set_state(ClientState.one_time_jobs_2)


@dp.message(ClientState.one_time_jobs_2)
async def one_time_jobs(message: Message, state: FSMContext) -> None:
    one_time_jobs_str = message.text.lower().split(', ')
    await edit_database(one_time_jobs=one_time_jobs_str)
    # with open('daily_scores.txt', 'r+', encoding='utf-8-sig') as f:
    #     data = json.load(f)
    #     data[str(message.from_user.id)]['one_time_jobs'] = one_time_jobs_str
    #     write_to_file(data, f)
    await message.answer('Отлично, теперь ваш список разовых дел выглядит так:')
    await message.answer(', '.join(one_time_jobs_str))
    await state.update_data(one_time_jobs=one_time_jobs_str)
    await settings(message, state)


@dp.message(F.text == 'Ежедневные дела', ClientState.jobs)
async def daily_jobs(message: Message, state: FSMContext) -> None:
    await message.answer('Введите новый список ежедневных. Ваш предыдущий список: ',
                         reply_markup=types.ReplyKeyboardRemove())
    user_states_data = await state.get_data()
    daily_scores = user_states_data['daily_scores']
    formatted_string = ""
    for key, value in daily_scores.items():
        formatted_string += f"{key}, "
    await message.answer(formatted_string[:-2])
    await message.answer(
        'Вы можете воспользоваться предложенным списком или написать свой. Данные могут быть какие '
        'угодно, очки нужны для отчетности о том насколько продуктивен был день.\nСоблюдайте формат '
        'данных!')
    await state.set_state(ClientState.new_daily_scores)


@dp.message(F.text == 'Скачать дневник')
async def download(message: Message) -> None:
    if os.path.exists(f'{message.from_user.username}_Diary.xlsx'):
        await message.answer_document(
            document=FSInputFile(f'{message.from_user.username}_Diary.xlsx'),
            disable_content_type_detection=True,
        )
    else:
        await message.answer('Сначала заполните данные')


@dp.message(ClientState.new_daily_scores)
async def command_start(message: Message, state: FSMContext):
    daily_scores = dict()
    user_message = message.text
    str_data = user_message.split(', ')
    for one_jobs in str_data:
        one_jobs = one_jobs.lower().replace('ё', 'е').split(' : ')
        if len(one_jobs) == 1:
            daily_scores[one_jobs[0]] = 1
        else:
            daily_scores[one_jobs[0]] = one_jobs[1]
    await state.update_data(daily_scores=daily_scores)
    await edit_database(daily_scores=daily_scores)
    # with open('daily_scores.txt', 'r+', encoding='utf-8-sig') as f:
    #     data = json.load(f)
    #     data[str(message.from_user.id)]['daily_scores'] = daily_scores
    #     write_to_file(data, f)
    await message.answer('Отлично, ваш список ежедневных дел обновлен!')
    await state.set_state(ClientState.settings)
    await settings(message, state)


@dp.message(ClientState.greet)
async def my_steps(message: Message, state: FSMContext):
    user_states_data = await state.get_data()
    daily_scores = user_states_data['daily_scores']
    # обработка ежедневных дел
    errors = [activity for activity in message.text.split(', ') if
              activity.lower().replace('ё', 'е') not in daily_scores]
    if errors:
        for error in errors:
            await message.answer(f"{error} нету в списке!")
    else:
        activities = [activity.lower().replace('ё', 'е') for activity in message.text.split(', ')]
        await state.update_data(activities=activities)
        if user_states_data['one_time_jobs'] != '':
            one_time_jobs = user_states_data['one_time_jobs']
            await message.answer('Введите разовые дела, которые выполнили. Список разовых дел:')
            await message.answer(', '.join(one_time_jobs))
            await state.set_state(ClientState.one_time_jobs_proceed)
        else:
            await message.answer("Сколько сделал шагов?")
            await state.set_state(ClientState.steps)


@dp.message(ClientState.one_time_jobs_proceed)
async def process_one_time(message: Message, state: FSMContext):
    text = message.text.split(', ')
    data = await state.get_data()
    one_time_jobs = data['one_time_jobs']
    for jobs in text:
        if jobs.lower().replace('ё', 'е') in ['не', 'нет', '-', 'pass', 'пасс', 'не хочу', 'скип',
                                             'пососи', 'пошел нахуй', 'неа', 'не-а', 0]:
            await message.answer("Сколько сделал шагов?")
            await state.set_state(ClientState.steps)
            return
        elif jobs.lower().replace('ё', 'е') not in one_time_jobs:
            await message.answer(f'{jobs} нету в списке разовых дел!')
            return
        else:
            one_time_jobs.remove(jobs.lower().replace('ё', 'е'))
    await edit_database(one_time_jobs=one_time_jobs)
    await state.update_data(one_time_jobs=one_time_jobs)
    if one_time_jobs == []:
        await message.answer('Поздравляю! Вы выполнили все разовые дела, так держать')
    await message.answer("Сколько сделал шагов?")
    await state.set_state(ClientState.steps)


@dp.message(ClientState.steps)
async def process_name(message: Message, state: FSMContext):
    await state.update_data(mysteps=message.text)
    await message.answer('Сколько всего спал?')
    await state.set_state(ClientState.total_sleep)


@dp.message(ClientState.total_sleep)
async def process_dont_like_write_bots(message: Message, state: FSMContext):
    await state.update_data(total_sleep=message.text)
    await message.answer('Сколько из них глубокий сон?')
    await state.set_state(ClientState.deep_sleep)


@dp.message(ClientState.deep_sleep)
async def process_like_write_bots(message: Message, state: FSMContext):
    await state.update_data(deep_sleep=message.text)
    await message.answer('Хочешь рассказать как прошел день? Это поможет отслеживать почему день был хороший или нет')
    await state.set_state(ClientState.about_day)


@dp.message(ClientState.about_day)
async def process_unknown_write_bots(message: Message, state: FSMContext):
    await state.update_data(user_message=message.text)
    await message.answer('Насколько из 10 сам оцениваешь день?')
    await state.set_state(ClientState.personal_rate)


@dp.message(ClientState.personal_rate)
async def process_unknown_write_bots(message: Message, state: FSMContext):
    user_states_data = await state.get_data()
    daily_scores = user_states_data['daily_scores']
    date = datetime.now()
    activities = user_states_data['activities']
    user_message = user_states_data['user_message']
    total_sleep = user_states_data['total_sleep']
    deep_sleep = user_states_data['deep_sleep']
    rate = message.text
    mysteps = user_states_data['mysteps']
    user_name = message.from_user.username
    await add_day_to_excel(date, activities, user_message, total_sleep, deep_sleep, rate, mysteps, message,
                           daily_scores,
                           user_name)
    await state.set_state(ClientState.greet)


async def add_day_to_excel(date, activities, user_message, total_sleep, deep_sleep, rate, mysteps, message,
                           daily_scores,
                           user_name):
    # Загрузка существующего файла Excel
    # with open('daily_scores.txt', 'r+', encoding='utf-8-sig') as f:
    #     line = json.load(f)
    #     for user in line:
    #         if message.from_user.id == user['user_id']:
    #             daily_scores = user['daily_scores']
    if os.path.exists(f'{user_name}_Diary.xlsx'):
        # If the file exists, load the existing workbook
        wb = load_workbook(f'{user_name}_Diary.xlsx')
    else:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1).value = 'Дата'
        sheet.cell(row=1, column=2).value = 'Дела за день'
        sheet.cell(row=1, column=3).value = 'Шаги'
        sheet.cell(row=1, column=4).value = 'Total sleep'
        sheet.cell(row=1, column=5).value = 'Deep sleep'
        sheet.cell(row=1, column=6).value = 'О дне'
        sheet.cell(row=1, column=7).value = 'My rate'
        sheet.cell(row=1, column=8).value = 'Total'

    # Получение активного листа
    sheet = wb.active

    # Поиск последней заполненной строки в столбце "Дата"
    last_row = sheet.max_row

    # Получение вчерашней даты
    yesterday = date - timedelta(days=1)

    # Вставка вчерашней даты в следующую за последней заполненной строкой
    sheet.cell(row=last_row + 1, column=1).value = yesterday.strftime("%d.%m.%Y")

    # Вставка дел в таблицу
    if type(activities) == list:
        sheet.cell(row=last_row + 1, column=2).value = ", ".join(activities)
    else:
        sheet.cell(row=last_row + 1, column=2).value = activities
    # Вставка шаги
    sheet.cell(row=last_row + 1, column=3).value = mysteps
    # Вставка сколько часов спал
    sheet.cell(row=last_row + 1, column=4).value = total_sleep
    # Вставка времени глубокого сна
    sheet.cell(row=last_row + 1, column=5).value = deep_sleep
    # Вставка о дне в таблицу
    sheet.cell(row=last_row + 1, column=6).value = user_message
    # Вставка времени глубокого сна
    sheet.cell(row=last_row + 1, column=7).value = rate

    # Вычисление итогового счета
    score = 0
    if type(activities) == list:
        for activity in activities:
            score += int(daily_scores[activity])
    else:
        score += int(daily_scores[activities])

    # Вставка итогового счета в таблицу
    sheet.cell(row=last_row + 1, column=8).value = score

    # Сохранение изменений в файл
    wb.save(f'{user_name}_Diary.xlsx')
    kb = [[types.KeyboardButton(text="Скачать дневник"), types.KeyboardButton(text="Изменить Настройки"),
           types.KeyboardButton(text="Вывести дневник")]]
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
    )

    async def counter_negavite(current_word):
        data = pd.read_excel(f'{user_name}_Diary.xlsx')

        # Выбор нужной колонки
        column = data['Дела за день']

        # Переменные для подсчета повторений
        count = 0
        # Проход по каждой строке колонки в обратном порядке
        for words in column.iloc[-1::-1]:
            flag = False
            for word in words.split(', '):
                if word == current_word:
                    # Если текущее слово совпадает с предыдущим,
                    # увеличиваем счетчик повторений
                    flag = True
            if not flag:
                count += 1
            if flag:
                return count
        return count

    async def counter_positive(current_word):
        data = pd.read_excel(f'{user_name}_Diary.xlsx')

        # Выбор нужной колонки
        column = data['Дела за день']

        # Переменные для подсчета повторений
        count = 0
        # Проход по каждой строке колонки в обратном порядке
        for words in column.iloc[-1::-1]:
            flag = False
            for word in words.split(', '):
                if word == current_word:
                    # Если текущее слово совпадает с предыдущим,
                    # увеличиваем счетчик повторений
                    count += 1
                    flag = True
            if not flag:
                if not count and count >= 2:
                    return count
                break
        if count >= 2:
            return count

    total_days = dict()
    for current_word in daily_scores:
        counter_days = await counter_negavite(current_word)
        if counter_days is not None and counter_days >= 3:
            if counter_days in [2, 3, 4]:
                total_days[current_word] = str(counter_days) + ' дня'
            else:
                total_days[current_word] = str(counter_days) + ' дней'
    output = ['{}: {}'.format(key, value) for key, value in total_days.items()]
    result = '\n'.join(output)
    if result != '':
        await message.answer(f'Вы не делали эти дела уже столько дней:\n{result}\nМожет стоит дать им еще один шанс?')
    total_days = dict()
    for current_word in activities:
        counter_days = await counter_positive(current_word)
        if counter_days in [2, 3, 4] and counter_days is not None:
            total_days[current_word] = str(counter_days) + ' дня'
        elif counter_days is not None:
            total_days[current_word] = str(counter_days) + ' дней'
    output = ['{}: {}'.format(key, value) for key, value in total_days.items()]
    result = '\n'.join(output)
    if result != '':
        await message.answer('Поздравляю! Вы соблюдаете эти дела уже столько дней: ' + '\n' + result,
                             reply_markup=keyboard)
    else:
        await message.answer('Дневник заполнен!', reply_markup=keyboard)


def write_to_file(data, f):
    f.truncate(0)
    f.seek(0)
    json.dump(data, f, ensure_ascii=False, indent=4)


async def existing_user(message, state):
    user_data = await state.get_data()
    if 'daily_scores' in user_data:
        daily_scores = user_data['daily_scores']
        await message.answer(
            'Расскажи мне как провел вчерашний день?' + '\n' + 'Вот список ежедневных дел:')
        await message.answer(', '.join(daily_scores.keys()), reply_markup=create_main_keyboard())
        await message.answer(
            'Впишите ежедневные дела которые вы вчера делали' + '\n'
            + 'Вы можете изменить списки в любой момент')
    if 'one_time_jobs' in user_data:
        one_time_jobs = user_data['one_time_jobs']
        await message.answer(
            'Разовые дела:')

        await message.answer(', '.join(one_time_jobs))
    await state.set_state(ClientState.greet)


def create_main_keyboard():
    kb = [[types.KeyboardButton(text="Скачать дневник"),
           types.KeyboardButton(text="Изменить Настройки"),
           types.KeyboardButton(text="Вывести дневник")]]
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
    )
    return keyboard


def create_change_settings_keyboard():
    kb = [[types.KeyboardButton(text="Изменить Настройки")]]
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
    )
    return keyboard


def create_settings_keyboard():
    kb = [[types.KeyboardButton(text="Изменить Дела"),
           types.KeyboardButton(text='''Что такое "стоимость"?'''), types.KeyboardButton(text="Изменить Стоимость"),
           types.KeyboardButton(text="Заполнить дневник")]]
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
    )
    return keyboard


def create_jobs_keyboard():
    kb = [[types.KeyboardButton(text="Ежедневные дела"),
           types.KeyboardButton(text="Разовые дела"),
           types.KeyboardButton(text='В определенную дату')]]
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
    )
    return keyboard


async def handle_new_user(message: Message, state):
    info = await bot.get_me()
    # with open('daily_scores.txt', 'w') as f:
    #     json.dump({int(message.from_user.id): {'daily_scores': [], 'one_time_jobs': [], 'date_jobs': {}}}, f,
    #               ensure_ascii=False, indent=4)
    markup = types.ReplyKeyboardRemove()
    await message.answer_sticker('CAACAgIAAxkBAAIsZGVY5wgzBq6lUUSgcSYTt99JnOBbAAIIAAPANk8Tb2wmC94am2kzBA')
    await message.answer(
        f'''Привет, {message.from_user.full_name}! \nДобро пожаловать в {info.username}!
Он поможет тебе вести отчет о твоих днях и делать выводы почему день был плохим или хорошим
Для начала нужно задать список дел. Какие у вас есть дела в течении дня? Например:''')
    await message.answer('встал в 6:30, лег в 11, зарядка утром, массаж, пп')
    await message.answer(
        'Вы можете воспользоваться предложенным списком или написать свой. Данные могут быть какие угодно',
        reply_markup=markup)
    await state.set_state(ClientState.new_daily_scores)


def start_scheduler(message, state):
    pass
    # if not hasattr(start_scheduler, "called"):
    #     scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
    #     scheduler.add_jobs(greetings, 'cron', hour=8, minute=45, args=(message, state))
    #     scheduler.start()


async def diary_out(message):
    data = pd.read_excel(f'{message.from_user.username}_Diary.xlsx')
    await message.answer(
        "{} | {} | {} | {} | {} | {} | {} | {}".format("Дата", "Дела за день", "Шаги", "Total sleep", 'Deep sleep',
                                                       'О дне', 'My rate', 'Total'))
    for index, row in data.iterrows():
        message_sheet = "{} | {} | {} | {} | {} | {} | {} | {}".format(row["Дата"], row["Дела за день"], row["Шаги"],
                                                                       row["Total sleep"], row['Deep sleep'],
                                                                       row['О дне'],
                                                                       row['My rate'], row['Total'])

        # разделение длинного сообщения на части
        message_parts = [message_sheet[i:i + 4096] for i in range(0, len(message_sheet), 4096)]

        for part in message_parts:
            await message.answer(part)


async def main():
    dp.startup.register(on_startup)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
